import os
import time
import cv2
import numpy as np
from matplotlib import pyplot as plt
from itertools import islice
import random
import re
from openpyxl import load_workbook
from pymouse import PyMouse
from pykeyboard import PyKeyboard
from PIL import Image, ImageGrab
from colorama import init
from termcolor import colored



def curl_get():
    print ('正在获取随机题库')
    for n in range(1,100):#循环下载20次随机题库，确保题目包含在内
        os.system('curl http://xxjs.dtdjzx.gov.cn/quiz-api/subject_info/randomList >> A:/randomList')
        n = n + 1

def get_screenshot():
    x_dim, y_dim = m.screen_size()
    img_rgb = ImageGrab.grab((0,0,x_dim*0.75,y_dim-40))
    img_rgb.save(sc_path)

def u_login():
    letter = ['uid']
    for choose_letter in letter:
        #time.sleep(1)#增加延时防止卷动屏幕时截图
        get_screenshot()
        img_rgb = cv2.imread(sc_path)
        img_gray = cv2.cvtColor(img_rgb, cv2.COLOR_BGR2GRAY)
        #print (choose_letter)
        template = cv2.imread('%s.png' % (choose_letter),0)
        w, h = template.shape[::-1]
        res = cv2.matchTemplate(img_gray,template,cv2.TM_CCOEFF_NORMED)
        threshold = 0.9
        loc = np.where( res >= threshold)
        for pt in zip(*loc[::-1]):
            cv2.rectangle(img_rgb, pt, (pt[0] + w, pt[1] + h), (0,0,255), 2)   
        print ('选项%s所在区域为：' % (choose_letter), pt[0] + w // 2, pt[1] + h // 2)
        x = int(pt[0] + w // 2) + 150 #确定uid窗口
        y = int(pt[1] + h // 2)
        m.click(x, y, 1)
        k.press_keys([k.control_l_key,'a']) #发送ctrl+a组合键并删除原uid
        k.tap_key(k.backspace_key)
        k.type_string(str(user))
        k.tap_key(k.tab_key)
        k.type_string(str(pswd))

def show_answer():
    f = open('A:/xxjs.dtdjzx.gov.cn/quiz-api/game_info/getGameSubject', 'r', encoding='utf-8') #正式答题路径
    #f = open('A:/xxjs.dtdjzx.gov.cn/quiz-api/subject_info/randomList', 'r', encoding='utf-8') #模拟答题路径
    a = f.read().replace('subjectTitle','\nsubjectTitle')  #处理文本，使其从固定地方断行
    #包含正式答题标题的正则pat_a
    pat_a = re.compile(r'subjectTitle":"(.*)","subjectType":"(.)","status":null,"answer":null,"totalRight":null,"totalWrong":null,"difficultyLevel":null,"createTime":null,"examFlag":null,"subjectId":null,"optionInfoList":\[{"id":"(.*)","optionTitle":"(.*)","optionType":"(.)"},{"id":"(.*)","optionTitle":"(.*)","optionType":"(.)"},{"id":"(.*)","optionTitle":"(.*)","optionType":"(.)"},{"id":"(.*)","optionTitle":"(.*)","optionType":"(.)"}\]')
    #pat_a = re.compile(r'subjectTitle":"(.*)","subjectType":"(.)","status')
    print ('正在答题')
    for index,line in enumerate(islice(a.splitlines(), 1, None)): #islice跳过读取第一行，依次循环读取
        print('=====第%s题=====' % (index + 1))
        print (pat_a.match(line).group(1))#1,4,5,7,8,10,11,13,14正则表达式的各个子组号 1为标题组
        subject = pat_a.match(line).group(1)
        f = open('A:/randomList', 'r', encoding='utf-8')
        b = f.read().replace('subjectTitle','\nsubjectTitle') #标题断行首字母subject
        pat_b = re.compile(r'subjectTitle":"(.*)","subjectType":"(.)","status":null,"answer":"(.*)","totalRight')
        dic = {}
        for line in islice(b.splitlines(), 1, None):
            dic[pat_b.match(line).group(1)] = pat_b.match(line).group(3).replace(',','') #生成标题对应答案的字典
        with open('dic.txt') as c: #插入不存在的答题到字典，文件在txt中，方便及时修改
            pat_c = re.compile(r'(.*)@(.*)@(.*)')
            for i in c.readlines():
                dic[pat_c.match(i).group(1)] = pat_c.match(i).group(2)
        if subject in dic:
            print ('++++++++++++字典中存在该答题')
        else:            
            print (colored(subject, 'white','on_red'))
            dic.setdefault(subject, 'a') #插入一个不存在的答题键值到字典，防止出现KeyError
        fin_answer = dic[subject]
        print ('>>>答案：{}'.format(dic[subject]))
        get_screenshot()
        img_rgb = cv2.imread(sc_path)
        img_gray = cv2.cvtColor(img_rgb, cv2.COLOR_BGR2GRAY)
        for  choose_letter in fin_answer+'n':#遍历答案寻找坐标
            template = cv2.imread('%s.png' % (choose_letter),0)
            w, h = template.shape[::-1]
            res = cv2.matchTemplate(img_gray,template,cv2.TM_CCOEFF_NORMED)
            threshold = 0.9
            loc = np.where( res >= threshold)
            for pt in zip(*loc[::-1]):
                cv2.rectangle(img_rgb, pt, (pt[0] + w, pt[1] + h), (0,0,255), 2)
            x = int(pt[0] + w // 2) #取整数并转化为整数型
            y = int(pt[1] + h // 2)
            if choose_letter == 'n':
                print ('>点击下一题坐标：{},{}'.format(x, y))
                m.click(x, y, 1)
            else:
                print ('>>>正在点击坐标: {} {},{}'.format(choose_letter, x, y))
                m.click(x, y, 1)
            time.sleep(random.randint(0, 10) / 4)#产生一个随机的时间延迟点击
        if index + 1 == 20:
            print ('>>>>>完成{}道答题'.format(index + 1))
        f.close()
#主函数
curl_get()
init() #初始化colorma颜色函数
m = PyMouse()
k = PyKeyboard()
sc_path = 'A:/autojump.png'
print ('>>当前屏幕分辨率为{}'.format(m.screen_size()))
print ('>>>开始答题')
wb = load_workbook('user.xlsx')
sheet = wb['Sheet1']
#print ('共计{}列'.format(sheet.max_column))#显示最大列
print ('共计{}人'.format(sheet.max_row))#显示最大行
for index,i in enumerate(range(1, sheet.max_row+1)):
    user = sheet.cell(row=i, column=1).value#i1单元格
    pswd = sheet.cell(row=i, column=2).value#i2单元格
    u_login()
    print ('当前第{}个答题用户为：{}，密码为{}' .format(index+1, user, pswd))
    input(">>>回车进行自动答题")
    '''for notime in range(3):
        input(">>>回车进行第{}次答题".format(notime))
        show_answer()#分享答题
    '''
    show_answer()
    os.system('echo {0}答题完毕 >> ./finish_uid.txt' .format(str(user).replace('\n', '')))#完成用户答题归档
    if index+1 == sheet.max_row:
        print ('所有人员答题完毕，共计{}人'.format(sheet.max_row))
    else:
        print ('用户{}答题完毕，准备继续作答，注销当前用户并返回到登录界面>>>'.format(user))
        input("按回车键继续...")
